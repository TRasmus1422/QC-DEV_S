'''
Created on 27. sep. 2023

@author: rn
'''

import os
import pandas as pd
import numpy as np
from openpyxl.chart import Reference, Series, ScatterChart
from openpyxl.chart.marker import Marker
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt

def startP1Datahandling(data,pathPic, writer):
    
    df_results, writer, fig1, fig2 = getData(data,writer,pathPic)
    
    return df_results, writer, fig1, fig2
    
def getData(data,writer,pathPic):
    
    #df = pd.read_excel(inputPath)

    #Pathsave = os.path.join(FolderPath, data["Roll ID"][0] + ".xlsx")
    
    df = createThresholdLines(data)
    
    df.to_excel(writer, sheet_name="p1_view")
    
    df_results = assesment(df, writer)
    
    df_results.to_excel(writer, sheet_name="Results")
    
    addAdjusted(df, writer)
    
    # P1 Original
    createScatter(df, writer, TR0=10, TR45=13, header = "P1 Original", location = "Y2")
    fig1 = create_scatter_show(df, TR0=10, TR45=1, header="P1 Original")
    
    # P1 Adjusted
    createScatter(df, writer, TR0=22, TR45=23, header = "P1 Adjusted", location = "BC2")
    #fig2 = create_scatter_show(df, TR0=22, TR45=23, header="P1 Adjusted")
    fig2 = None
    
    #findPicture(df, writer, pathPic)
    
    return df_results, writer, fig1, fig2

def create_scatter_show(df, TR0=None, TR45=None, header = "None"):
    fig, ax = plt.subplots()
    ax.scatter(df['RollPos (m)'], df.iloc[:, TR0], color='blue', label='TR 0')
    ax.scatter(df['RollPos (m)'], df.iloc[:, TR45], color='red', label='TR 45')

    # Adding flat threshold lines
    thresholds = [80, 76, 44, 38]
    colors = ['black', 'black', 'black', 'black']  # Colors for the threshold lines
    for threshold, color in zip(thresholds, colors):
        ax.hlines(y=threshold, xmin=df['RollPos (m)'].min(), xmax=df['RollPos (m)'].max(), colors=color, linestyles='dashed', label=f'Threshold at {threshold}')

    ax.set_title(header)
    ax.set_xlabel('RollPos (m)')
    ax.set_ylabel('Transmittans[%]')
    ax.legend()

    return fig

def findPicture(df, writer, p1Path):
    
    workbook = writer.book
    ws = workbook.create_sheet("pics")
    #ws = workbook["pics"]
    
    TR0PICS = df[(df["TR 0 Sharp"] > 70) & ((df["TR 0"] < 76) | (df["TR 0"] > 80))]

    TR45PICS = df[(df["TR 45 Sharp"] > 80) & ((df["TR 45"] > 44) | (df["TR 45"] < 38))]
    
    TR0_PICS = TR0PICS["PICTURE NAME 0"]
    TR0_ROOL = TR0PICS["RollPos (m)"].tolist()
    TR0_SHARP = TR0PICS["TR 0 Sharp"].tolist()
    TR0_VALUE = TR0PICS["TR 0"].tolist()
    TR0_DEV = TR0PICS["TR 0 DEV"].tolist()
    
    
    TR45_PICS = TR45PICS["PICTURE NAME 45"].tolist()
    TR45_ROOL = TR45PICS["RollPos (m)"].tolist()
    TR45_SHARP = TR45PICS["TR 45 Sharp"].tolist()
    TR45_VALUE = TR45PICS["TR 45"].tolist()
    TR45_DEV = TR45PICS["TR 45 DEV"].tolist()
    
    roll_ID = df["Roll ID"][0]

    startNum = 2

    for i, pic in enumerate(TR0_PICS):
        
        img_id_0 = pic.replace('"', '') + ".jpg"
        path_0 = os.path.join(p1Path, roll_ID + "_Log", "NG-billeder", img_id_0)
        
        my_png = Image(path_0)
        my_png.width = 512
        my_png.height = 260.5
        ws[f'B{startNum+(i*20)-1}'] = f'Cam 0 - RollPos (m) : {str(TR0_ROOL[i])} - TR0 Sharp : {str(TR0_SHARP[i])} - TR0 : {str(TR0_VALUE[i])} - TR0 DEV : {str(TR0_DEV[i])}'
        ws.add_image(my_png, f'B{startNum+i*20}')
        
    for i, pic in enumerate(TR45_PICS):
        
        img_id_45 = pic.replace('"', '') + ".jpg"
        path_45 = os.path.join(p1Path, roll_ID + "_Log", "NG-billeder", img_id_45)
        
        my_png = Image(path_45)
        my_png.width = 512
        my_png.height = 260.5
        ws[f'K{startNum+(i*20)-1}'] = f'Cam 45 - RollPos (m) : {str(TR45_ROOL[i])} - TR45 Sharp : {str(TR45_SHARP[i])} - TR45 : {str(TR45_VALUE[i])} - TR45 DEV : {str(TR45_DEV[i])}'
        ws.add_image(my_png, f'K{startNum+i*20}')
    
def createThresholdLines(df):
    
    length_data = len(df["RollPos (m)"])
    
    top0 = np.full(length_data, 80)
    bottom0 = np.full(length_data, 76)
    top45 = np.full(length_data, 44)
    bottom45 = np.full(length_data, 38)
    
    df['top0'] = top0.tolist()
    df['bottom0'] = bottom0.tolist()
    df['top45'] = top45.tolist()
    df['bottom45'] = bottom45.tolist()
    
    return df

def addAdjusted(df, writer):
    
    workbook = writer.book
    ws = workbook["p1_view"]
    
    ws['U1'] = "Adjust Value"
    ws['U2'] = 2
    
    ws['V1'] = "Adjusted TR 0"
    ws['W1'] = "Adjusted TR 45"
    
    # Write a formula into cell B1 that sums up A1 and A2
    for i in range(len(df["TR 0"])):
        ws[f'V{i+2}'] = f"=J{i+2}+U2"
        ws[f'W{i+2}'] = f"=M{i+2}+U2"
    
def assesment(df,writer):
    
    dict = {"Description":["Good", "Bad", "N/A", "Total", "Precentage"],
            "Sharpness 0":[],
            "% in spec UVT0":[],
            "Sharpness 45":[],
            "% in spec UVT45":[]
            }
    
    ## 0 Angle
    
    good_sharpness_0 = len(df[(df["TR 0 Sharp"] >= 75)])
    dict["Sharpness 0"].append(good_sharpness_0)
    
    bad_sharpness_0 = len(df[(df["TR 0 Sharp"] < 75) & (df["TR 0 Sharp"] > 0)])
    dict["Sharpness 0"].append(bad_sharpness_0)
    
    good_inSpec_0 = len(df[(df["TR 0 Sharp"] >= 75) & (df["TR 0"] <= 80) & (df["TR 0"] >= 76)])
    dict["% in spec UVT0"].append(good_inSpec_0)
    
    bad_inSpec_0 = len(df[(df["TR 0 Sharp"] >= 75) & (df["TR 0"] < 76)]) + len(df[(df["TR 0 Sharp"] >= 75) & (df["TR 0"] > 80)])
    dict["% in spec UVT0"].append(bad_inSpec_0)
    
    NA_inSpec_0 = len(df[(df["TR 0"] == 0)])
    dict["% in spec UVT0"].append(NA_inSpec_0)
    
    total_0 = good_inSpec_0 + bad_inSpec_0 - NA_inSpec_0
    dict["% in spec UVT0"].append(total_0)
    
    precentage_0 = good_inSpec_0 / total_0 * 100
    dict["% in spec UVT0"].append(precentage_0)
    
    sharpness_0 = good_sharpness_0 / (good_sharpness_0+bad_sharpness_0)*100
    dict["Sharpness 0"].append(None)
    dict["Sharpness 0"].append(None)
    dict["Sharpness 0"].append(sharpness_0)
    
    ## 45 Angle
    
    good_sharpness_45 = len(df[(df["TR 45 Sharp"] >= 75)])
    dict["Sharpness 45"].append(good_sharpness_45)
    
    bad_sharpness_45 = len(df[(df["TR 45 Sharp"] < 75) & (df["TR 45 Sharp"] > 0)])
    dict["Sharpness 45"].append(bad_sharpness_45)
    
    good_inSpec_45 = len(df[(df["TR 45 Sharp"] >= 75) & (df["TR 45"] <= 44) & (df["TR 45"] >= 38)])
    dict["% in spec UVT45"].append(good_inSpec_45)
    
    bad_inSpec_45 = len(df[(df["TR 45 Sharp"] >= 75) & (df["TR 45"] < 38)]) + len(df[(df["TR 45 Sharp"] >= 75) & (df["TR 45"] > 44)])
    dict["% in spec UVT45"].append(bad_inSpec_45)
    
    NA_inSpec_45 = len(df[(df["TR 45"] == 0)])
    dict["% in spec UVT45"].append(NA_inSpec_45)

    total_45 = good_inSpec_45 + bad_inSpec_45 - NA_inSpec_45
    dict["% in spec UVT45"].append(total_45)
    
    precentage_45 = good_inSpec_45 / total_45 * 100
    dict["% in spec UVT45"].append(precentage_45)
    
    sharpness_45 = good_sharpness_45 / (good_sharpness_45+bad_sharpness_45)*100
    dict["Sharpness 45"].append(None)
    dict["Sharpness 45"].append(None)
    dict["Sharpness 45"].append(sharpness_45)
    
    df_results = pd.DataFrame.from_dict(dict)
    
    return df_results

def createScatter(df, writer, TR0=10, TR45=13, header = "P1 Original", location = "Y2"):
    
    workbook = writer.book
    ws = workbook["p1_view"]
    
    chart = ScatterChart()
    chart.title = header
    chart.x_axis.title = 'RollPos (m)'
    chart.y_axis.title = 'Transmittans[%]'
    chart.width = 50  # Adjust as needed
    chart.height = 20  # Adjust as needed
    
    length_data = len(df["RollPos (m)"])
    

    ## Threshold Lines
    
    Xvalues = Reference(ws, min_col=7, min_row=2, max_row=length_data)
    threshold_lines=[]
    threshold_lines.append(Reference(ws, min_col=17, min_row=2, max_row=length_data))
    threshold_lines.append(Reference(ws, min_col=18, min_row=2, max_row=length_data))
    threshold_lines.append(Reference(ws, min_col=19, min_row=2, max_row=length_data))
    threshold_lines.append(Reference(ws, min_col=20, min_row=2, max_row=length_data))
    
    for threshold in threshold_lines:
        series2 = Series(threshold, Xvalues, title="Threshold")
        series2.graphicalProperties.line.noFill = False
        chart.series.append(series2)
        
    ## Values    
    values = []
    values.append(Reference(ws, min_col=TR0, min_row=2, max_row=length_data))
    values.append(Reference(ws, min_col=TR45, min_row=2, max_row=length_data))
    
    title = ["TR 0","TR 45"]
    
    for i in range(len(values)):
        series = Series(values[i], Xvalues, title=title[i])
        
        marker = Marker('circle')
        marker.size = 3  # Adjust size as needed
        series.marker = marker
        
        series.graphicalProperties.line.noFill = True
        chart.series.append(series)
    
    ws.add_chart(chart, location)

if __name__ == '__main__':
    startP1Datahandling()